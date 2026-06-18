import openpyxl

file_path = "d:/OCRProject/output/Test13_Centrifugal_Pump_extracted.xlsm"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb.active

target_pages = [4, 5, 13, 40]
print("--- ACTUAL OUTPUT FROM TEST 13 SCRIPT ---")
count = 0
for row_idx in range(3, ws.max_row + 1):
    page = ws.cell(row=row_idx, column=13).value
    if page in target_pages:
        pos = ws.cell(row=row_idx, column=8).value
        part = ws.cell(row=row_idx, column=6).value
        name = ws.cell(row=row_idx, column=5).value
        qty = ws.cell(row=row_idx, column=12).value
        sub = ws.cell(row=row_idx, column=2).value
        print(f"Page {page} | Pos: {pos} | Part: {part} | Name: {name} | Qty: {qty} | Sub: {sub}")
        count += 1

print(f"Total rows in actual output for target pages: {count}")
