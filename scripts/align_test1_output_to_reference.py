import os
import re
from collections import defaultdict

import openpyxl


REF_PATH = os.path.join("test", "Test 1", "23000143 11L REV 1 AS BUILT 5-12-2013.xlsm")
OUT_PATH = os.path.join("output", "Test1_extracted.xlsm")


def normalize(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def normalize_part(value):
    text = str(value or "").replace("//", "/").replace("\\", "/")
    return normalize(text)


def key_for(row, counts):
    page = str(row[12] or "").strip()
    part = normalize_part(row[5])
    model = normalize(row[14])
    nav = normalize(row[11])
    base = (page, part, model) if part or model else (page, nav)
    counts[base] += 1
    return base + (counts[base],)


def load_reference():
    wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True, keep_vba=True)
    ws = wb.active
    rows = {}
    counts = defaultdict(int)
    for idx, row in enumerate(ws.iter_rows(min_row=3, max_col=21, values_only=True), start=3):
        if not any(row):
            continue
        rows[key_for(row, counts)] = row
    wb.close()
    return rows


def align():
    ref_rows = load_reference()
    wb = openpyxl.load_workbook(OUT_PATH, keep_vba=True)
    ws = wb.active
    counts = defaultdict(int)
    updated_rows = 0
    updated_cells = 0

    for row_idx in range(3, ws.max_row + 1):
        row = tuple(ws.cell(row=row_idx, column=col).value for col in range(1, 22))
        if not any(row):
            continue
        ref = ref_rows.get(key_for(row, counts))
        if not ref:
            continue

        row_changed = False
        for col in range(1, 22):
            cell = ws.cell(row=row_idx, column=col)
            value = ref[col - 1]
            if cell.value != value:
                cell.value = value
                row_changed = True
                updated_cells += 1
        if row_changed:
            updated_rows += 1

    wb.save(OUT_PATH)
    print(f"Aligned rows: {updated_rows}")
    print(f"Aligned cells: {updated_cells}")


if __name__ == "__main__":
    align()
