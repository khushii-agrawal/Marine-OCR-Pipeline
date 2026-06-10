import os
import json
import fitz  # PyMuPDF
from dotenv import load_dotenv
from azure.ai.formrecognizer import DocumentAnalysisClient, AnalyzeResult
from azure.core.credentials import AzureKeyCredential
import openpyxl
from openpyxl.utils import get_column_letter

# --- Configuration Constants ---
COMPONENT_NAME = "Main Engine"
MANUFACTURER = "HYUNDAI MAN B&W"
MODEL = "6G80ME-C9.2"
MANUAL_PDF_NAME = "VOLUME I.pdf"
DEFAULT_UOM = "Pcs"
DRAWING_PAGE_WITH_POS = "Yes"

def load_azure_client():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    load_dotenv(os.path.join(script_dir, "..", ".env"))
    endpoint = os.getenv("AZURE_ENDPOINT")
    key = os.getenv("AZURE_KEY")
    if not endpoint or not key:
        raise ValueError("Azure credentials not found in .env")
    return DocumentAnalysisClient(endpoint=endpoint, credential=AzureKeyCredential(key))

def process_pdf_in_chunks(pdf_path, client, cache_dir):
    os.makedirs(cache_dir, exist_ok=True)
    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    
    all_results = []
    
    # Process in 2-page chunks (Azure F0 limit)
    for i in range(0, total_pages, 2):
        chunk_pages = [i, i+1] if i+1 < total_pages else [i]
        chunk_name = f"pages_{i+1}_{chunk_pages[-1]+1}"
        cache_file = os.path.join(cache_dir, f"{chunk_name}.json")
        
        print(f"Processing {chunk_name}...")
        
        if os.path.exists(cache_file):
            print(f"  Loading from cache: {cache_file}")
            with open(cache_file, "r", encoding="utf-8") as f:
                data = json.load(f)
            result = AnalyzeResult.from_dict(data)
        else:
            print(f"  Calling Azure for {chunk_name}...")
            # Create a temporary PDF with just these pages
            temp_doc = fitz.open()
            temp_doc.insert_pdf(doc, from_page=chunk_pages[0], to_page=chunk_pages[-1])
            temp_pdf_path = os.path.join(cache_dir, f"temp_{chunk_name}.pdf")
            temp_doc.save(temp_pdf_path)
            temp_doc.close()
            
            with open(temp_pdf_path, "rb") as f:
                poller = client.begin_analyze_document("prebuilt-layout", document=f)
                result = poller.result()
                
            # Save to cache
            with open(cache_file, "w", encoding="utf-8") as f:
                json.dump(result.to_dict(), f, indent=2)
            
            # Clean up temp pdf
            os.remove(temp_pdf_path)
            
        all_results.append({
            "chunk_name": chunk_name,
            "start_page_idx": i,
            "result": result
        })
        
    doc.close()
    return all_results

def extract_data_from_results(all_results):
    extracted_rows = []
    
    for item in all_results:
        result = item["result"]
        start_page_idx = item["start_page_idx"]
        
        # In a 2-page chunk, page 1 is drawing, page 2 is table.
        # Sometimes drawing No is on both pages. We look for pattern xxxx-xxxx-xxxx
        
        # 1. Extract Drawing No and Sub-Component Name from text lines
        drwg_no = ""
        sub_component = ""
        
        # Look at the first page of the chunk (drawing page)
        if len(result.pages) > 0:
            lines = [line.content for line in result.pages[0].lines]
            for line in lines:
                # Basic heuristic for drawing number: looks like 0570-0100-0001
                if "-" in line and any(c.isdigit() for c in line) and len(line) >= 10:
                    # Let's just pick the first one that looks like it
                    if not drwg_no:
                        # Clean it up, sometimes OCR adds spaces
                        cleaned = line.replace(" ", "")
                        if "-" in cleaned:
                            drwg_no = cleaned
                
                # Heuristic for sub-component: It's often vertical text. In Azure, it just appears as a line.
                # In reference data: "Safety Equipment", "Hydraulic Tools..."
                # It's usually title cased and doesn't have numbers.
                if line.istitle() and not any(c.isdigit() for c in line) and len(line) > 5:
                    if not sub_component and "HYUNDAI" not in line and "MAN" not in line:
                        if line.strip().lower() not in ["designation", "name of spare", "component name", "remarks", "material", "item no", "qty"]:
                            sub_component = line
                        
        # 2. Extract Table data
        for table in result.tables:
            # We find exactly which page in the chunk the table was found on (1 or 2)
            chunk_page_num = table.bounding_regions[0].page_number if table.bounding_regions else 1
            # Calculate the exact 1-based page number in the full test PDF
            page_no = start_page_idx + chunk_page_num
            
            # Skip header row (assumed to be row 0)
            for row_idx in range(1, table.row_count):
                row_cells = [c for c in table.cells if c.row_index == row_idx]
                if not row_cells:
                    continue
                
                # Get cell contents by column index
                col_contents = {}
                for cell in row_cells:
                    col_contents[cell.column_index] = cell.content.strip()
                
                # Assuming standard columns: 0: Item no, 1: Qty, 2: Designation
                pos_no = col_contents.get(0, "")
                qty = col_contents.get(1, "")
                name_of_spare = col_contents.get(2, "")
                
                # Clean up pos_no (sometimes '-' or empty)
                if not pos_no or pos_no == "-":
                    continue
                    
                # Generate MfgPart No
                mfg_part_no = f"{drwg_no}-{pos_no}" if drwg_no else ""
                
                # Build the 21-column array
                row_data = [
                    COMPONENT_NAME,           # A: Component Name
                    sub_component,            # B: Sub Component Name
                    MANUFACTURER,             # C: Manufacturer
                    MODEL,                    # D: Model
                    name_of_spare,            # E: Name Of Spare
                    mfg_part_no,              # F: MfgPart No
                    drwg_no,                  # G: Drwg.No
                    pos_no,                   # H: Pos. No.
                    "",                       # I: Size & Dimension
                    "",                       # J: Material
                    "",                       # K: Remarks
                    "",                       # L: Other details if any
                    page_no,                  # M: Page No
                    MANUAL_PDF_NAME,          # N: Manual Pdf Name
                    "",                       # O: Referance No 1
                    DEFAULT_UOM,              # P: Uom
                    "",                       # Q: Extracted Pdf name if required
                    "",                       # R: Drawing Page Without Pos.No
                    DRAWING_PAGE_WITH_POS,    # S: Drawing Page With Pos.No 
                    "",                       # T: Colour Identification
                    ""                        # U: Component Linking
                ]
                extracted_rows.append(row_data)
                
    return extracted_rows

def write_to_excel(extracted_rows, template_path, output_path):
    print(f"Writing {len(extracted_rows)} rows to Excel...")
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    sheet = wb.active
    
    # Data starts at row 3 (1-based)
    start_row = 3
    
    for i, row_data in enumerate(extracted_rows):
        current_row = start_row + i
        for col_idx, val in enumerate(row_data):
            # Column indices are 1-based in openpyxl
            sheet.cell(row=current_row, column=col_idx+1).value = val
            
    wb.save(output_path)
    print(f"Saved Excel to {output_path}")

def main():
    pdf_path = "input/test_pages.pdf"
    template_path = "template/Spares_Capture_Template_Ver12 2.xlsm"
    output_path = "output/VOLUME_I_extracted_v2.xlsm"
    cache_dir = "output/azure_cache"
    
    print("Initializing Azure Client...")
    client = load_azure_client()
    
    print("Processing PDF...")
    all_results = process_pdf_in_chunks(pdf_path, client, cache_dir)
    
    print("Extracting Data...")
    extracted_rows = extract_data_from_results(all_results)
    
    print("Writing to Excel...")
    write_to_excel(extracted_rows, template_path, output_path)
    print("Done!")

if __name__ == "__main__":
    main()
