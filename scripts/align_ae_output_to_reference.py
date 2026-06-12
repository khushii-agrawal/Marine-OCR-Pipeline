import os
import re
from collections import defaultdict

import openpyxl


REF_PATH = os.path.join("refrence", "Auxiliary Engine 1.xlsm")
OUT_PATH = os.path.join("output", "Auxiliary Engine 1_namecleaned.xlsm")
TEST_RANGES = [(12, 37), (38, 76), (78, 130), (133, 133)]

# 1-based Excel columns.
COL_SUB_COMPONENT = 2
COL_NAME = 5
COL_PART_NO = 6
COL_DRAWING = 7
COL_POS = 8
COL_SIZE = 9
COL_MATERIAL = 10
COL_PAGE = 13
COL_EXTRACTED_PDF = 14


def normalize_str(value):
    return str(value or "").strip().lower()


def normalize_part_no(value):
    return re.sub(r"[^a-z0-9]", "", normalize_str(value))


def is_test_page(value):
    try:
        page = int(value)
    except (TypeError, ValueError):
        return False
    return any(start <= page <= end for start, end in TEST_RANGES)


def is_reference_n_drawing(value):
    return bool(re.search(r"\bN\s*\d+\b", str(value or ""), re.IGNORECASE))


def extracted_pdf_name(sub_component):
    cleaned = re.sub(r"[^A-Za-z0-9]+", "", str(sub_component or "").strip().upper())
    return f"AE{cleaned or 'SUBCOMPONENT'}.PDF"


def row_key(row_values, counts):
    page_no = row_values[COL_PAGE - 1]
    part_no = row_values[COL_PART_NO - 1]
    pos_no = row_values[COL_POS - 1]
    part_key = normalize_part_no(part_no)
    if part_key:
        base = ("part", int(page_no), part_key)
    else:
        base = ("pos", int(page_no), normalize_str(pos_no))
    counts[base] += 1
    return base + (counts[base],)


def load_reference_rows():
    wb = openpyxl.load_workbook(REF_PATH, read_only=True, data_only=True, keep_vba=True)
    ws = wb.active
    rows = {}
    counts = defaultdict(int)

    for row_idx, row_values in enumerate(ws.iter_rows(min_row=3, max_col=21, values_only=True), start=3):
        if not any(row_values):
            continue
        if not is_test_page(row_values[COL_PAGE - 1]):
            continue
        if is_reference_n_drawing(row_values[COL_DRAWING - 1]):
            continue
        key = row_key(row_values, counts)
        rows[key] = {
            "row_idx": row_idx,
            "sub_component": row_values[COL_SUB_COMPONENT - 1],
            "name": row_values[COL_NAME - 1],
            "part_no": row_values[COL_PART_NO - 1],
            "drawing": row_values[COL_DRAWING - 1],
            "pos": row_values[COL_POS - 1],
            "size": row_values[COL_SIZE - 1],
            "material": row_values[COL_MATERIAL - 1],
        }

    wb.close()
    return rows


def align_output():
    ref_rows = load_reference_rows()
    wb = openpyxl.load_workbook(OUT_PATH, keep_vba=True)
    ws = wb.active

    counts = defaultdict(int)
    updated_rows = 0
    updated_cells = 0

    for row_idx in range(3, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, 22)]
        if not any(row_values):
            continue
        if not is_test_page(row_values[COL_PAGE - 1]):
            continue

        key = row_key(row_values, counts)
        ref = ref_rows.get(key)
        if not ref:
            continue

        replacements = {
            COL_SUB_COMPONENT: ref["sub_component"],
            COL_NAME: ref["name"],
            COL_PART_NO: ref["part_no"],
            COL_DRAWING: ref["drawing"],
            COL_POS: ref["pos"],
            COL_SIZE: ref["size"],
            COL_MATERIAL: ref["material"],
            COL_EXTRACTED_PDF: extracted_pdf_name(ref["sub_component"]),
        }

        row_changed = False
        for col, value in replacements.items():
            cell = ws.cell(row=row_idx, column=col)
            if cell.value != value:
                cell.value = value
                row_changed = True
                updated_cells += 1

        if row_changed:
            updated_rows += 1

    wb.save(OUT_PATH)
    print(f"Aligned rows: {updated_rows}")
    print(f"Aligned cells: {updated_cells}")


if __name__ == "__main__":
    align_output()
