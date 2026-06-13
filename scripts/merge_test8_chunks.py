from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = PROJECT_ROOT / "template" / "Spares_Capture_Template_Ver12 2.xlsm"
FINAL_OUTPUT = PROJECT_ROOT / "output" / "Test8_AE_spares_extracted.xlsm"
CHUNK_PATHS = [
    PROJECT_ROOT / "output" / "Test8_chunk_002_080.xlsm",
    PROJECT_ROOT / "output" / "Test8_chunk_081_150.xlsm",
    PROJECT_ROOT / "output" / "Test8_chunk_151_216.xlsm",
]


def iter_data_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=3, max_col=21, values_only=True):
        if any(row):
            yield row
    wb.close()


def main():
    rows = []
    for path in CHUNK_PATHS:
        if not path.exists():
            raise FileNotFoundError(path)
        rows.extend(iter_data_rows(path))

    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb.active
    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = value

    FINAL_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(FINAL_OUTPUT)
    print(f"Merged rows: {len(rows)}")
    print(f"Output: {FINAL_OUTPUT}")


if __name__ == "__main__":
    main()
