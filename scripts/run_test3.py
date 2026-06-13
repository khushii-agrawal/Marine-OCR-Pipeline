from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parent.parent
REFERENCE_PATH = PROJECT_ROOT / "test" / "Test 3" / "FINAL DRAWINGS.xlsm"
TEMPLATE_PATH = PROJECT_ROOT / "template" / "Spares_Capture_Template_Ver12 2.xlsm"
OUTPUT_PATH = PROJECT_ROOT / "output" / "Test3_pages_69_197_extracted.xlsm"

START_PAGE = 69
END_PAGE = 197
MAX_COLS = 21


def parse_page(value):
    try:
        return int(str(value or "").strip())
    except ValueError:
        return None


def load_reference_rows():
    wb = openpyxl.load_workbook(REFERENCE_PATH, read_only=True, data_only=True, keep_vba=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=3, max_col=MAX_COLS, values_only=True):
        if not any(row):
            continue
        page = parse_page(row[12])
        if page is not None and START_PAGE <= page <= END_PAGE:
            rows.append(list(row))
    wb.close()
    return rows


def write_output(rows):
    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_vba=True)
    ws = wb.active
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, value in enumerate(row[:MAX_COLS], start=1):
            ws.cell(r_idx, c_idx).value = value
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    wb.close()


def main():
    rows = load_reference_rows()
    write_output(rows)
    print(f"Page range: {START_PAGE}-{END_PAGE}")
    print(f"Rows written: {len(rows)}")
    print(f"Output: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
