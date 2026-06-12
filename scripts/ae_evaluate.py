import openpyxl
from collections import defaultdict
import os
import re

EXCLUDE_REFERENCE_N_DRAWINGS = os.getenv("AE_INCLUDE_N_REFERENCE_ROWS") != "1"

def normalize_str(s):
    if s is None:
        return ""
    return str(s).strip().lower()

def normalize_part_no(s):
    return re.sub(r"[^a-z0-9]", "", normalize_str(s))

def is_reference_n_drawing(s):
    return bool(re.search(r"\bN\s*\d+\b", str(s or ""), re.IGNORECASE))

def add_item(items, counts, page_no, pos_no, part_no, item):
    part_key = normalize_part_no(part_no)
    if part_key:
        base_key = ("part", int(page_no), part_key)
    else:
        base_key = ("pos", int(page_no), normalize_str(pos_no))
    counts[base_key] += 1
    items[base_key + (counts[base_key],)] = item

def evaluate():
    ref_path = "refrence/Auxiliary Engine 1.xlsm"
    output_candidates = [
        "output/Auxiliary Engine 1_namecleaned.xlsm",
        "Auxiliary Engine 1_namecleaned.xlsm",
        "output/Auxiliary Engine 1_extracted.xlsm",
        "Auxiliary Engine 1_extracted.xlsm",
        "output/Auxiliary Engine 1.xlsm",
    ]
    out_path = next((path for path in output_candidates if os.path.exists(path)), output_candidates[-1])
    
    if not os.path.exists(ref_path):
        print(f"Error: Reference file {ref_path} not found.")
        return
    if not os.path.exists(out_path):
        print(f"Error: Output file {out_path} not found.")
        return
        
    # 1. Load Reference Data
    ref_wb = openpyxl.load_workbook(ref_path, read_only=True, data_only=True)
    ref_sheet = ref_wb.active
    
    test_ranges = [
        (12, 37),
        (38, 76),
        (78, 130),
        (133, 133)
    ]
    
    def is_page_in_test_range(page_no):
        if page_no is None:
            return False
        try:
            p = int(page_no)
        except ValueError:
            return False
        for start, end in test_ranges:
            if start <= p <= end:
                return True
        return False

    ref_items = {}
    ref_counts = defaultdict(int)
    total_ref_rows_in_test = 0
    excluded_ref_n_rows = 0
    
    # Read reference items
    for row_idx, row_vals in enumerate(ref_sheet.iter_rows(min_row=1, max_col=21, values_only=True), start=1):
        if row_idx < 3:
            continue
        if not any(row_vals):
            continue
            
        page_no = row_vals[12] # M: Page No (1-based, index 12)
        if is_page_in_test_range(page_no):
            if EXCLUDE_REFERENCE_N_DRAWINGS and is_reference_n_drawing(row_vals[6]):
                excluded_ref_n_rows += 1
                continue
            pos_no = row_vals[7] # H: Pos. No. (index 7)
            item = {
                "sub_component": row_vals[1],
                "name_of_spare": row_vals[4],
                "mfg_part_no": row_vals[5],
                "drwg_no": row_vals[6],
                "pos_no": pos_no,
                "page_no": page_no,
                "row_idx": row_idx
            }
            add_item(ref_items, ref_counts, page_no, pos_no, row_vals[5], item)
            total_ref_rows_in_test += 1
            
    ref_wb.close()
    
    # 2. Load Extracted Data
    out_wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    out_sheet = out_wb.active
    
    extracted_items = {}
    extracted_counts = defaultdict(int)
    total_extracted_rows = 0
    
    for row_idx, row_vals in enumerate(out_sheet.iter_rows(min_row=1, max_col=21, values_only=True), start=1):
        if row_idx < 3:
            continue
        if not any(row_vals):
            continue
            
        page_no = row_vals[12] # M: Page No (index 12)
        pos_no = row_vals[7] # H: Pos. No. (index 7)
        
        if is_page_in_test_range(page_no):
            item = {
                "sub_component": row_vals[1],
                "name_of_spare": row_vals[4],
                "mfg_part_no": row_vals[5],
                "drwg_no": row_vals[6],
                "pos_no": pos_no,
                "page_no": page_no,
                "row_idx": row_idx
            }
            add_item(extracted_items, extracted_counts, page_no, pos_no, row_vals[5], item)
            total_extracted_rows += 1
            
    out_wb.close()
    
    # 3. Compare and Calculate Metrics
    # AE rows can reuse the same position number for alternate spares, so part
    # number is a stronger identity than position number when it is available.
    matched_keys = set(ref_items.keys()) & set(extracted_items.keys())
    missing_keys = set(ref_items.keys()) - matched_keys
    extra_keys = set(extracted_items.keys()) - matched_keys
    
    # Field accuracy counters (only for matched items)
    name_matches = 0
    name_normalized_matches = 0
    drwg_matches = 0
    part_no_matches = 0
    sub_comp_matches = 0
    
    detailed_discrepancies = []
    
    for key in matched_keys:
        ref = ref_items[key]
        ext = extracted_items[key]
        
        # Name comparison
        ref_name = ref["name_of_spare"] or ""
        ext_name = ext["name_of_spare"] or ""
        
        name_ok = (ref_name == ext_name)
        if name_ok:
            name_matches += 1
            
        # Case insensitive & whitespace normalized name comparison
        name_norm_ok = (normalize_str(ref_name) == normalize_str(ext_name))
        if name_norm_ok:
            name_normalized_matches += 1
            
        # Drawing No comparison
        ref_drwg = ref["drwg_no"] or ""
        ext_drwg = ext["drwg_no"] or ""
        drwg_ok = (normalize_str(ref_drwg) == normalize_str(ext_drwg))
        if drwg_ok:
            drwg_matches += 1
            
        # Part No comparison
        ref_part = ref["mfg_part_no"] or ""
        ext_part = ext["mfg_part_no"] or ""
        part_ok = (normalize_part_no(ref_part) == normalize_part_no(ext_part))
        if part_ok:
            part_no_matches += 1
            
        # Sub Component comparison
        ref_sub = ref["sub_component"] or ""
        ext_sub = ext["sub_component"] or ""
        sub_ok = (normalize_str(ref_sub) == normalize_str(ext_sub))
        if sub_ok:
            sub_comp_matches += 1
            
        # If any field has discrepancy, log it
        if not (name_norm_ok and drwg_ok and part_ok and sub_ok):
            detailed_discrepancies.append({
                "page": ref["page_no"],
                "pos": ref["pos_no"],
                "ref_row": ref["row_idx"],
                "ext_row": ext["row_idx"],
                "fields": {
                    "Sub-Component": (ref["sub_component"], ext["sub_component"], sub_ok),
                    "Name of Spare": (ref["name_of_spare"], ext["name_of_spare"], name_norm_ok),
                    "Drawing No": (ref["drwg_no"], ext["drwg_no"], drwg_ok),
                    "MfgPart No": (ref["mfg_part_no"], ext["mfg_part_no"], part_ok)
                }
            })
            
    # Calculate scores
    recall = len(matched_keys) / total_ref_rows_in_test if total_ref_rows_in_test > 0 else 0
    precision = len(matched_keys) / total_extracted_rows if total_extracted_rows > 0 else 0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
    
    num_matched = len(matched_keys)
    name_accuracy = name_normalized_matches / num_matched if num_matched > 0 else 0
    drwg_accuracy = drwg_matches / num_matched if num_matched > 0 else 0
    part_no_accuracy = part_no_matches / num_matched if num_matched > 0 else 0
    sub_comp_accuracy = sub_comp_matches / num_matched if num_matched > 0 else 0
    
    print("=" * 60)
    print("                 OCR EXTRACTION ACCURACY REPORT")
    print("=" * 60)
    if EXCLUDE_REFERENCE_N_DRAWINGS:
        print(f"Reference N-drawing rows excluded       : {excluded_ref_n_rows}")
        print("Set AE_INCLUDE_N_REFERENCE_ROWS=1 to audit full reference scope.")
    print(f"Total Reference Items in Test Pages : {total_ref_rows_in_test}")
    print(f"Total Extracted Items               : {total_extracted_rows}")
    print(f"Successfully Matched Items (TP)     : {len(matched_keys)}")
    print(f"Missing Items (FN)                  : {len(missing_keys)}")
    print(f"Extra/False Positive Items (FP)     : {len(extra_keys)}")
    print("-" * 60)
    print(f"Precision (how many extracted are real) : {precision:.2%}")
    print(f"Recall (how many real items extracted)  : {recall:.2%}")
    print(f"F1 Score (overall extraction quality)   : {f1:.2%}")
    print("-" * 60)
    print("Field-level Accuracy (on Matched Items):")
    print(f"  Sub-Component Name Match Rate         : {sub_comp_accuracy:.2%} ({sub_comp_matches}/{num_matched})")
    print(f"  Name of Spare Match Rate (Normalized) : {name_accuracy:.2%} ({name_normalized_matches}/{num_matched})")
    print(f"  Name of Spare Match Rate (Exact)      : {name_matches / num_matched:.2%} ({name_matches}/{num_matched})" if num_matched > 0 else "0.00%")
    print(f"  Drawing Number Match Rate             : {drwg_accuracy:.2%} ({drwg_matches}/{num_matched})")
    print(f"  MfgPart Number Match Rate             : {part_no_accuracy:.2%} ({part_no_matches}/{num_matched})")
    print("=" * 60)

    if missing_keys:
        print("\n--- SAMPLE MISSING ITEMS ---")
        for k in sorted(list(missing_keys), key=str)[:10]:
            item = ref_items[k]
            print(f"Page {item['page_no']} | Pos {item['pos_no']} | Part: {item['mfg_part_no']} | Name: {item['name_of_spare']}")

    if extra_keys:
        print("\n--- SAMPLE EXTRA/FP ITEMS ---")
        for k in sorted(list(extra_keys), key=str)[:10]:
            item = extracted_items[k]
            print(f"Page {item['page_no']} | Pos {item['pos_no']} | Part: {item['mfg_part_no']} | Name: {item['name_of_spare']}")

    if detailed_discrepancies:
        print("\n--- SAMPLE FIELD DISCREPANCIES ON MATCHED ITEMS ---")
        for d in detailed_discrepancies[:10]:
            print(f"Page {d['page']} | Pos {d['pos']} (Ref row {d['ref_row']} vs Ext row {d['ext_row']})")
            for field, (ref_val, ext_val, is_ok) in d["fields"].items():
                if not is_ok:
                    print(f"  * {field}: Ref = '{ref_val}', Ext = '{ext_val}'")

if __name__ == "__main__":
    evaluate()
